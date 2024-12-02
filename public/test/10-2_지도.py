import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import re

filePath = '고등교육기관 하반기 주소록(2023).xlsx'
df_from_excel = pd.read_excel(filePath, engine='openpyxl')
df_from_excel.columns = df_from_excel.loc[4].tolist()
df_from_excel = df_from_excel.drop(index=list(range(0, 5)))


url = 'http://api.vworld.kr/req/address?'
params = 'service=address&request=getcoord&version=2.0&crs=epsg:4326&refine=true&simple=false&format=json&type='
road_type = 'ROAD'  # 도로명 주소
road_type2 = 'PARCEL'  # 지번 주소
address = '&address='
keys = '&key='
primary_key = 'D0A6060F-D988-36D4-B7AA-0649CF0C8011'


def request_geo(road):
    page = requests.get(url+params+road_type+address+road+keys+primary_key)
    json_data = page.json()  # 실행한 결과를 json형태로 받아옴

    # 응답의 상태가 ok인 경우 -> 요청에 대해 응답을 정상적으로 받았으면
    if json_data['response']['status'] == 'OK':
        x = json_data['response']['result']['point']['x']
        y = json_data['response']['result']['point']['y']
        return x, y
    else:
        x = 0
        y = 0
        return x, y

try:
    # 학교주소좌표.xlsx파일을 열고 시트를 활성화 시키는데
    # 만약 해당 파일이 없으면 예외가 발생
    wb = load_workbook('학교주소좌표.xlsx', data_only=True)
    sheet = wb.active
except:
    # 새로운 파일을 생성하고 시트를 활성화
    wb = Workbook()
    sheet = wb.active

# 자료에서 학교명과 주소를 리스트 형태로 변환하여 저장
university_list = df_from_excel['학교명'].to_list()
address_list = df_from_excel['주소'].to_list()

# enumerate : 반복 -> 자바에선 이터레이터
for num, value in enumerate(address_list):
    
    # 주소에서 괄호와 괄호안의 모든 문자열을 제거 , ^표시는 제외하는거 : ^) 은 )을 빼고 찾아라
    addr = re.sub(r'\(    [^)]*   \)', '', value)
    # print(addr)
    x,y = request_geo(addr)
    sheet.append([university_list[num], addr, x, y])

wb.save('학교주소좌표.xlsx')
